import os
from configparser import ConfigParser

import logging
import pandas as pd
from Updated_rows import find_updated_rows  # Import the logic
from attribut_comparison import compare_all_headers
from enodecategory_comparison import compare_enodecategory
import json


def setup_logging(log_file_path=r"C:\Accuris_Kapow\TE_COMPARISON REPORT\Comparison_report.log"):
    """
    Sets up logging for the script.
    """
    logging.basicConfig(
        level=logging.INFO,
        format="%(asctime)s - %(levelname)s - %(message)s",
        handlers=[
            logging.FileHandler(log_file_path),
            logging.StreamHandler()
        ]
    )


def convert_to_parquet(input_file, parquet_file, chunk_size=10000):
    """
    Processes an input file (Excel/CSV) in chunks and writes to Parquet format.

    Parameters:
        input_file (str): Path to the input file.
        parquet_file (str): Path to the output Parquet file.
        chunk_size (int): Number of rows per chunk to process.

    Returns:
        str: Path to the output Parquet file.
    """
    logging.info(f"Starting chunk-wise processing of {input_file}.")
    print(f"Starting chunk-wise processing of {input_file}.")

    rows_processed = 0
    chunks = []  # List to hold data chunks

    # Reader logic for Excel/CSV files
    if input_file.endswith('.xlsx') or input_file.endswith('.xls'):
        from openpyxl import load_workbook

        wb = load_workbook(input_file, read_only=True)
        ws = wb.active
        headers = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]

        def excel_reader():
            chunk = []
            for row in ws.iter_rows(min_row=2, values_only=True):
                chunk.append(row)
                if len(chunk) == chunk_size:
                    yield pd.DataFrame(chunk, columns=headers)
                    chunk = []
            if chunk:
                yield pd.DataFrame(chunk, columns=headers)

        reader = excel_reader()
    elif input_file.endswith('.csv'):
        reader = pd.read_csv(input_file, dtype=str, chunksize=chunk_size, encoding='utf-8')
    else:
        raise ValueError("Unsupported file format. Only Excel and CSV are supported.")

    # Process each chunk
    for chunk_index, chunk in enumerate(reader):
        rows_processed += len(chunk)
        logging.info(f"Processing chunk {chunk_index + 1}: {len(chunk)} rows.")
        print(f"Processing chunk {chunk_index + 1}: {len(chunk)} rows.")
        chunks.append(chunk)

    # Concatenate all chunks into a single DataFrame
    logging.info("Combining all chunks into a single DataFrame.")
    print("Combining all chunks into a single DataFrame.")
    full_df = pd.concat(chunks, ignore_index=True)

    # Write the combined DataFrame to Parquet
    logging.info(f"Writing final DataFrame to Parquet: {parquet_file}")
    print(f"Writing final DataFrame to Parquet: {parquet_file}")
    full_df.to_parquet(parquet_file, index=False, engine='pyarrow')

    logging.info(f"Completed processing of {input_file}. Total rows processed: {rows_processed}.")
    print(f"Chunk-wise processing completed: {rows_processed} rows processed.")
    return parquet_file


def compare_excel_files(file1, file2, output_file_path, part_reference_column, category_comparison_columns,
                        attribute_reference_column):
    """
    Compares two Excel files and generates a detailed report.
    """
    # Log input details for debugging
    print(f"Comparing files:\n  File 1: {file1}\n  File 2: {file2}\n  Output: {output_file_path}\n")
    # Read the Excel files with all columns as strings
    logging.info(f"Starting comparison of files:\n  File 1: {file1}\n  File 2: {file2}")
    print(f"Starting comparison of files:\n  File 1: {file1}\n  File 2: {file2}")

    # Read the files
    logging.info(f"Reading File 1: {file1}")
    df1 = pd.read_parquet(file1)
    logging.info(f"File 1 loaded with {len(df1)} rows.")

    logging.info(f"Reading File 2: {file2}")
    df2 = pd.read_parquet(file2)
    logging.info(f"File 2 loaded with {len(df2)} rows.")

    # Ensure 'TCPN' is treated as string to avoid matching issues
    df1[part_reference_column] = df1[part_reference_column].astype(
        str).str.strip() if part_reference_column in df1.columns else ""
    df2[part_reference_column] = df2[part_reference_column].astype(
        str).str.strip() if part_reference_column in df2.columns else ""

    # Set 'TCPN' as index for easier comparison
    if part_reference_column in df1.columns and part_reference_column in df2.columns:
        df1.set_index(part_reference_column, inplace=True)
        df2.set_index(part_reference_column, inplace=True)

    # Use the separated logic to find updated rows
    updated_rows = find_updated_rows(df1, df2)

    # Identify new rows
    new_rows = []
    for tcpn in df2.index.difference(df1.index):
        new_row = df2.loc[tcpn].to_dict()
        new_row[part_reference_column] = tcpn
        new_rows.append(new_row)

    # Find deleted rows
    deleted_rows = df1.loc[~df1.index.isin(df2.index)]
    deleted_rows = deleted_rows.reset_index()  # Reset index to include 'TCPN'

    attribute_comparison_results = compare_all_headers(df1, df2, reference_column=attribute_reference_column)

    enode_results = compare_enodecategory(df1, df2, reference_column=category_comparison_columns)

    summary_data = {
        'Description': [
            '# Parts data updated',
            '# New Parts',
            '# Deleted Parts',
            'Attribute_Details_No_Change',
            'Attribute_Details_New',
            'Attribute_Details_Deleted',
            'Attribute_Details_updated',
            'Category_Details_No_Change',
            'Category_Details_New',
            'Category_Details_updated',
            'Category_Details_Deleted'
        ],
        'Count': [
            len(updated_rows),
            len(new_rows),
            len(deleted_rows),
            len(attribute_comparison_results[attribute_comparison_results['Status'] == 'No Change']),
            len(attribute_comparison_results[attribute_comparison_results['Status'] == 'New']),
            len(attribute_comparison_results[attribute_comparison_results['Status'] == 'Deleted']),
            len(attribute_comparison_results[attribute_comparison_results['Status'] == 'Updated']),
            len(enode_results[enode_results['Status'] == 'No_Change']),
            len(enode_results[enode_results['Status'] == 'New']),
            len(enode_results[enode_results['Status'] == 'Updated']),
            len(enode_results[enode_results['Status'] == 'Deleted'])
        ]
    }
    summary_df = pd.DataFrame(summary_data)

    # Create a new Excel writer
    with pd.ExcelWriter(output_file, engine='xlsxwriter') as writer:
        summary_df.to_excel(writer, sheet_name='Summary', index=False)

        # Write updated rows to 'Updated Rows' sheet (only changed columns)
        if not updated_rows.empty:  # Check if there are updated rows
            updated_rows.to_excel(writer, sheet_name='Parts_data_updated', index=False)
            format_rows(writer, 'Parts_data_updated', 'green')

        # Write new rows to 'New Rows' sheet
        if new_rows:
            new_df = pd.DataFrame(new_rows)
            new_df = reorder_columns(new_df, part_reference_column)  # Ensure TCPN is the first column
            new_df.to_excel(writer, sheet_name='New_parts', index=False)
            format_rows(writer, 'New_parts', 'blue')

        # Write deleted rows to 'Deleted Rows' sheet
        if not deleted_rows.empty:
            deleted_df = reorder_columns(deleted_rows, part_reference_column)  # Ensure TCPN is the first column
            deleted_df.to_excel(writer, sheet_name='Deleted_parts', index=False)
            format_rows(writer, 'Deleted_parts', 'red')

        if not attribute_comparison_results.empty:
            attribute_comparison_results.to_excel(writer, sheet_name='Attribute_Details', index=False)
            format_rows(writer, 'Attribute_Details', 'black')

        if not enode_results.empty:
            enode_results.to_excel(writer, sheet_name='Category_Details', index=False)
            format_rows(writer, 'Category_Details', 'orange')

    print(f"Report generated: {output_file}")


# Rest of your helper functions remain unchanged


def reorder_columns(df, first_column):
    """
    Reorders the columns of the DataFrame to place `first_column` at the start.
    """
    columns = [first_column] + [col for col in df.columns if col != first_column]
    return df[columns]


def format_rows(writer, sheet_name, color):
    workbook = writer.book
    worksheet = writer.sheets[sheet_name]
    format_color = workbook.add_format({'font_color': color})
    worksheet.set_column('A:Z', 20)  # Adjust column width as necessary
    worksheet.set_row(0, None, workbook.add_format({'bold': True}))  # Bold header
    for row_num in range(1, worksheet.dim_rowmax + 1):  # Skip header
        worksheet.set_row(row_num, None, format_color)


if __name__ == "__main__":
    setup_logging()
    chunk_size = 10000
    # Path to the configuration file
    config_file = r"C:\Accuris_Kapow\TE_COMPARISON REPORT\Config.ini"

    # Check if the configuration file exists
    if not os.path.exists(config_file):
        raise FileNotFoundError(f"Configuration file not found: {config_file}")

    # Load the configuration
    config = ConfigParser()
    config.read(config_file)

    try:
        # Retrieve file paths and settings
        log_filepath = config['Paths']["log_filepath"]
        input_file_1 = config["Paths"]["input_file_1"]
        input_file_2 = config["Paths"]["input_file_2"]
        output_file = config["Paths"]["output_file"]
        reference_column = config["Settings"]["reference_column"]
        comparison_columns = config["Settings"]["category_comparison_columns"]
        attribute_reference_column = config["Settings"]["attribute_reference_column"]
        setup_logging(log_filepath)
        # Validate input files
        if not os.path.exists(input_file_1):
            raise FileNotFoundError(f"Input file 1 not found: {input_file_1}")
        if not os.path.exists(input_file_2):
            raise FileNotFoundError(f"Input file 2 not found: {input_file_2}")

        parquet_file_1 = convert_to_parquet(input_file_1, "file1.parquet", chunk_size)
        parquet_file_2 = convert_to_parquet(input_file_2, "file2.parquet", chunk_size)

        # Compare files
        compare_excel_files(
            parquet_file_1,
            parquet_file_2,
            output_file,
            reference_column,
            comparison_columns,
            attribute_reference_column
        )

    except KeyError as e:
        raise KeyError(f"Missing configuration key: {e}")

    except Exception as e:
        print(f"An error occurred: {e}")
