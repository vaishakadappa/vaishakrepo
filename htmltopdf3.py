import openpyxl
from weasyprint import HTML

# Replace with your Excel file path
excel_file_path = 'C:\\vaishak\\TAIY_1.5K_28Sp.xlsx'

# Load the Excel workbook
workbook = openpyxl.load_workbook(excel_file_path)

# Define a function to convert a URL to PDF
def convert_url_to_pdf(url, output_pdf):
    try:
        html = HTML(url)
        html.write_pdf(output_pdf)
        print(f'Converted {url} to {output_pdf}')
    except Exception as e:
        print(f'Error converting {url} to PDF: {str(e)}')

# Iterate through all the sheets in the workbook
for sheet in workbook.sheetnames:
    worksheet = workbook[sheet]

    # Iterate through all cells to find links
    for row in worksheet.iter_rows(values_only=True):
        for cell_value in row:
            if cell_value is not None and 'http' in str(cell_value).lower():
                link_url = str(cell_value)
                pdf_output_path = f'output_{sheet}_{row[0].row}.pdf'

                # Convert the link to PDF
                convert_url_to_pdf(link_url, pdf_output_path)

# Save the changes to the Excel file if needed
# workbook.save(excel_file_path)

# Close the workbook
workbook.close()
